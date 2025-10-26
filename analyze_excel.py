#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
تحليل شامل لملف Excel الخاص بنظام CRM
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
from datetime import datetime

def analyze_excel_file(file_path):
    """تحليل شامل لملف Excel"""
    
    print("=" * 80)
    print("تحليل شامل لملف Excel - نظام CRM")
    print("=" * 80)
    print()
    
    # تحميل الملف
    wb = load_workbook(file_path, data_only=True)
    
    analysis = {
        'file_info': {},
        'sheets': {},
        'data_structure': {},
        'relationships': [],
        'business_logic': {},
        'statistics': {}
    }
    
    # معلومات عامة عن الملف
    analysis['file_info'] = {
        'file_name': file_path,
        'total_sheets': len(wb.sheetnames),
        'sheet_names': wb.sheetnames,
        'analysis_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    print(f"📁 اسم الملف: {file_path}")
    print(f"📊 عدد الأوراق: {len(wb.sheetnames)}")
    print(f"📋 أسماء الأوراق: {', '.join(wb.sheetnames)}")
    print()
    
    # تحليل كل ورقة
    for sheet_name in wb.sheetnames:
        print("=" * 80)
        print(f"📄 تحليل الورقة: {sheet_name}")
        print("=" * 80)
        
        # قراءة البيانات باستخدام pandas
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        ws = wb[sheet_name]
        
        # معلومات أساسية
        sheet_analysis = {
            'name': sheet_name,
            'dimensions': {
                'rows': len(df),
                'columns': len(df.columns),
                'total_cells': len(df) * len(df.columns)
            },
            'columns': [],
            'data_types': {},
            'sample_data': {},
            'statistics': {},
            'null_values': {},
            'unique_values': {}
        }
        
        print(f"\n📐 الأبعاد:")
        print(f"   - عدد الصفوف: {len(df)}")
        print(f"   - عدد الأعمدة: {len(df.columns)}")
        print(f"   - إجمالي الخلايا: {len(df) * len(df.columns)}")
        
        print(f"\n📋 الأعمدة ({len(df.columns)}):")
        for idx, col in enumerate(df.columns, 1):
            col_data = df[col]
            
            # تحليل العمود
            col_info = {
                'name': str(col),
                'index': idx,
                'data_type': str(col_data.dtype),
                'non_null_count': int(col_data.count()),
                'null_count': int(col_data.isna().sum()),
                'null_percentage': float(col_data.isna().sum() / len(df) * 100) if len(df) > 0 else 0,
                'unique_count': int(col_data.nunique()),
                'unique_percentage': float(col_data.nunique() / len(df) * 100) if len(df) > 0 else 0
            }
            
            # إضافة عينة من البيانات
            sample_values = col_data.dropna().head(5).tolist()
            col_info['sample_values'] = [str(v) for v in sample_values]
            
            # إحصائيات للأعمدة الرقمية
            if pd.api.types.is_numeric_dtype(col_data):
                col_info['statistics'] = {
                    'min': float(col_data.min()) if not col_data.isna().all() else None,
                    'max': float(col_data.max()) if not col_data.isna().all() else None,
                    'mean': float(col_data.mean()) if not col_data.isna().all() else None,
                    'median': float(col_data.median()) if not col_data.isna().all() else None,
                    'std': float(col_data.std()) if not col_data.isna().all() else None
                }
            
            # القيم الفريدة للأعمدة ذات القيم المحدودة
            if col_info['unique_count'] <= 20 and col_info['unique_count'] > 0:
                unique_vals = col_data.dropna().unique().tolist()
                col_info['all_unique_values'] = [str(v) for v in unique_vals]
            
            sheet_analysis['columns'].append(col_info)
            
            # طباعة معلومات العمود
            print(f"\n   {idx}. {col}")
            print(f"      - نوع البيانات: {col_info['data_type']}")
            print(f"      - القيم غير الفارغة: {col_info['non_null_count']}")
            print(f"      - القيم الفارغة: {col_info['null_count']} ({col_info['null_percentage']:.1f}%)")
            print(f"      - القيم الفريدة: {col_info['unique_count']} ({col_info['unique_percentage']:.1f}%)")
            
            if col_info['sample_values']:
                print(f"      - عينة من البيانات: {', '.join(col_info['sample_values'][:3])}")
            
            if 'statistics' in col_info:
                stats = col_info['statistics']
                if stats['min'] is not None:
                    print(f"      - الإحصائيات: Min={stats['min']:.2f}, Max={stats['max']:.2f}, Mean={stats['mean']:.2f}")
            
            if 'all_unique_values' in col_info:
                print(f"      - جميع القيم الفريدة: {', '.join(col_info['all_unique_values'])}")
        
        # عرض عينة من البيانات
        print(f"\n📊 عينة من البيانات (أول 5 صفوف):")
        print(df.head().to_string())
        
        # تحليل العلاقات المحتملة
        print(f"\n🔗 تحليل العلاقات المحتملة:")
        for col in df.columns:
            col_str = str(col).lower()
            if 'id' in col_str:
                print(f"   - {col}: محتمل أن يكون مفتاح أساسي أو خارجي")
            elif 'name' in col_str or 'اسم' in col_str:
                print(f"   - {col}: حقل اسم")
            elif 'date' in col_str or 'تاريخ' in col_str:
                print(f"   - {col}: حقل تاريخ")
            elif 'email' in col_str or 'بريد' in col_str:
                print(f"   - {col}: حقل بريد إلكتروني")
            elif 'phone' in col_str or 'هاتف' in col_str or 'جوال' in col_str:
                print(f"   - {col}: حقل هاتف")
            elif 'status' in col_str or 'حالة' in col_str:
                print(f"   - {col}: حقل حالة")
            elif 'price' in col_str or 'amount' in col_str or 'سعر' in col_str or 'مبلغ' in col_str:
                print(f"   - {col}: حقل مالي")
        
        analysis['sheets'][sheet_name] = sheet_analysis
        print()
    
    # حفظ التحليل في ملف JSON
    output_file = 'excel_analysis.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(analysis, f, ensure_ascii=False, indent=2)
    
    print("=" * 80)
    print(f"✅ تم حفظ التحليل الكامل في الملف: {output_file}")
    print("=" * 80)
    
    return analysis

if __name__ == "__main__":
    file_path = "CRM (1).xlsx"
    analysis = analyze_excel_file(file_path)

